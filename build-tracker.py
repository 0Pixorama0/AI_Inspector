# -*- coding: utf-8 -*-
"""Build the AI Device Inspector delivery tracker (.xlsx)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
import datetime

# ---- brand ----
VIOLET="7C3AED"; BLUE="3A33C9"; CYAN="0E9AAC"; MAGENTA="B23FC9"
INK="1A1B2E"; MUTED="6A6C83"; PAPER2="F4F5FA"; LINEC="D9DCEA"
G_DONE="D6F3E4"; G_PROG="D7E8FB"; G_BLOCK="FBD9D9"; G_REV="FCEFCD"; G_NEW="ECEDF3"
WHITE="FFFFFF"; YELLOW="FFF6CC"
FONT="Arial"

def F(sz=10,b=False,color=INK,it=False): return Font(name=FONT,size=sz,bold=b,color=color,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=LINEC)
box=Border(left=thin,right=thin,top=thin,bottom=thin)
bottom=Border(bottom=Side(style="thin",color=LINEC))
center=Alignment(horizontal="center",vertical="center")
left=Alignment(horizontal="left",vertical="center",wrap_text=True)
lefttop=Alignment(horizontal="left",vertical="top",wrap_text=True)

wb=openpyxl.Workbook()

# ============================================================ DATA
# (id, workstream, deliverable, task, sprint, priority, est_days)
tasks=[
 ("T-001","General","—","Repo scaffolding, monorepo & coding standards","Sprint 1","High",1),
 ("T-002","General","—","CI/CD pipeline (GitHub Actions) + SCA/SAST/secret scan","Sprint 1","High",2),
 ("T-003","General","—","Terraform: staging & production environments","Sprint 1","High",2),
 ("T-004","WS1","D-05","Design tokens + component library","Sprint 1","High",3),
 ("T-005","WS1","D-05","App shell, navigation & layouts (3 surfaces)","Sprint 1","High",2),
 ("T-006","WS1","D-01","Email/password auth + optional 2FA","Sprint 1","High",2),
 ("T-007","WS1","D-01","Enterprise SSO (SAML/OIDC) + rotating JWT","Sprint 1","High",2),
 ("T-008","WS1","D-02","Role model + permission matrix (server-side)","Sprint 1","High",2),
 ("T-009","WS1","D-02","Per-role permission test suite","Sprint 1","Medium",1),
 ("T-010","WS1","D-03","Postgres schema + migrations","Sprint 1","High",2),
 ("T-011","WS1","D-03","Row-level security (multi-tenant isolation)","Sprint 1","High",2),
 ("T-012","WS1","D-04","Station kiosk sessions (device tokens, idle lock, switch)","Sprint 1","Medium",2),
 ("T-013","WS2","D-06","Six-state guided capture workflow UI","Sprint 2","High",3),
 ("T-014","WS2","D-07","Barcode/QR scan + OCR service-tag reading","Sprint 2","High",2),
 ("T-015","WS2","D-07","Device lookup (model / spec / history / warranty)","Sprint 2","Medium",1),
 ("T-016","WS2","D-08","Camera controller integration (6-camera trigger)","Sprint 2","High",3),
 ("T-017","WS2","D-08","Focus/exposure validation + per-angle retake","Sprint 2","Medium",2),
 ("T-018","WS2","D-08","Presigned image uploads to object storage","Sprint 2","High",1),
 ("T-019","WS2","D-09","Offline capture queue + reconnect sync","Sprint 2","Medium",2),
 ("T-020","WS2","—","Live station status over WebSocket","Sprint 2","Medium",1),
 ("T-021","WS3","D-10","Defect taxonomy definition + sign-off","Sprint 2","High",1),
 ("T-022","WS3","—","Training-data labelling pipeline","Sprint 2","High",2),
 ("T-023","WS3","D-11","Fine-tune detection model (v1)","Sprint 2","High",3),
 ("T-024","WS3","D-12","Inference service (FastAPI + ONNX/TensorRT)","Sprint 2","High",3),
 ("T-025","WS3","D-12","Confidence thresholds + auto-escalation","Sprint 2","Medium",1),
 ("T-026","WS3","D-11","Detection storage (region / severity / heat map)","Sprint 2","High",2),
 ("T-027","WS3","—","Inference job queue (Redis + BullMQ)","Sprint 2","High",1),
 ("T-028","WS3","D-13","Model versioning + result stamping","Sprint 2","Medium",1),
 ("T-029","WS4","D-14","Annotated viewer (boxes, heat map, zoom-to-defect)","Sprint 3","High",3),
 ("T-030","WS4","D-14","Cursor tool: resize / draw defect regions","Sprint 3","High",2),
 ("T-031","WS4","D-15","Detection management (accept/reject/reclassify/merge)","Sprint 3","High",2),
 ("T-032","WS4","D-15","Low-confidence queue + supervisor routing","Sprint 3","Medium",1),
 ("T-033","WS4","D-16","Grading engine (cosmetic/functional + health score)","Sprint 3","High",3),
 ("T-034","WS4","D-16","Per-client grade profiles (versioned)","Sprint 3","High",2),
 ("T-035","WS4","D-17","Disposition rules + supervisor override w/ reason","Sprint 3","Medium",1),
 ("T-036","WS3","D-13","Retrain loop: corrections into training set","Sprint 3","Medium",2),
 ("T-037","General","—","Automated regression suite (build out)","Sprint 3","High",2),
 ("T-038","WS5","D-18","Inspection history grid (virtualised, 16-col)","Sprint 4","High",3),
 ("T-039","WS5","D-18","Search, filters, saved presets, bulk actions","Sprint 4","High",2),
 ("T-040","WS5","D-19","Inspection record tabs + full-res viewer","Sprint 4","High",2),
 ("T-041","WS5","D-19","Branded PDF certificate + tamper hash + share link","Sprint 4","High",2),
 ("T-042","WS5","D-20","CSV import wizard (5-step)","Sprint 4","Medium",2),
 ("T-043","WS5","D-20","Exports (CSV / XLSX / PDF / JSON)","Sprint 4","Medium",1),
 ("T-044","WS5","D-21","REST API + OpenAPI + signed webhooks","Sprint 4","High",2),
 ("T-045","WS5","D-21","ERP/WMS connector + ZPL label printing","Sprint 4","Medium",2),
 ("T-046","WS6","D-22","Dashboard & live floor (tiles, KPIs, charts)","Sprint 4","High",3),
 ("T-047","WS6","D-23","Analytics (throughput, grade dist, defect Pareto)","Sprint 4","Medium",2),
 ("T-048","WS6","D-24","Administration (clients/users/stations/taxonomy/profiles)","Sprint 4","High",3),
 ("T-049","WS6","D-25","Immutable audit log + export","Sprint 4","High",2),
 ("T-050","WS6","D-25","Retention/purge + right-to-erasure + redaction","Sprint 4","Medium",2),
 ("T-051","General","—","Accessibility pass (WCAG 2.2 AA)","Sprint 4","Medium",2),
 ("T-052","General","—","Performance budget enforcement (LCP / INP / p95)","Sprint 4","Medium",1),
 ("T-053","General","—","System integration testing (SIT)","Test","High",3),
 ("T-054","General","—","Regression suite green (3 consecutive runs)","Test","High",2),
 ("T-055","General","—","Data migration rehearsal + rollback drill","Test","Medium",1),
 ("T-056","WS3","—","Model accuracy validation on holdout set","Test","High",2),
 ("T-057","General","—","UAT support (operator / supervisor scripts)","Test","High",3),
 ("T-058","General","—","Load & soak testing (3x peak, 24h)","Test","Medium",2),
 ("T-059","General","—","Third-party penetration test + remediation","Test","High",3),
 ("T-060","General","—","Failure drills (network / GPU / storage)","Test","Medium",1),
 ("T-061","General","—","Documentation (architecture, API, guides, runbooks)","Test","Medium",3),
 ("T-062","General","—","Training + recorded walkthroughs","Test","Medium",1),
 ("T-063","General","—","Pilot line monitoring + go-live cutover","Go-live","High",3),
]
N=len(tasks); LAST=1+N  # data rows 2..LAST

STATUSES=["Not Started","In Progress","Blocked","In Review","Done"]
PRIOS=["High","Medium","Low"]
SPRINTS=["Sprint 1","Sprint 2","Sprint 3","Sprint 4","Test","Go-live"]
WSS=["WS1","WS2","WS3","WS4","WS5","WS6","General"]
WS_NAME={"WS1":"Foundations","WS2":"Capture","WS3":"AI engine","WS4":"Review & grading",
         "WS5":"Records & data","WS6":"Insight & admin","General":"Cross-cutting"}

# ============================================================ LISTS (hidden)
ls=wb.active; ls.title="Lists"
cols={"A":STATUSES,"B":PRIOS,"C":SPRINTS,"D":WSS}
heads={"A":"Status","B":"Priority","C":"Sprint","D":"Workstream"}
for col,vals in cols.items():
    ls[f"{col}1"]=heads[col]; ls[f"{col}1"].font=F(9,True,MUTED)
    for i,v in enumerate(vals): ls[f"{col}{i+2}"]=v
ls.sheet_state="hidden"

# ============================================================ TASK TRACKER
tt=wb.create_sheet("Task Tracker")
headers=["Task ID","Workstream","Deliverable","Task","Sprint","Priority","Assignee",
         "Est (days)","Start","Due","Status","% Complete","Notes / blockers"]
widths=[9,11,11,46,10,10,16,9,12,12,13,11,34]
for c,(h,w) in enumerate(zip(headers,widths),1):
    cell=tt.cell(1,c,h); cell.font=F(10,True,WHITE); cell.fill=fill(VIOLET)
    cell.alignment=center; cell.border=box
    tt.column_dimensions[get_column_letter(c)].width=w
tt.row_dimensions[1].height=26

for r,(tid,ws,dv,task,sp,pr,est) in enumerate(tasks,2):
    row=[tid,ws,dv,task,sp,pr,"",est,"","","Not Started",0,""]
    for c,val in enumerate(row,1):
        cell=tt.cell(r,c,val); cell.border=box; cell.font=F(10)
        if c in (1,2,3,5,6,8,11): cell.alignment=center
        elif c==12: cell.alignment=center
        else: cell.alignment=lefttop
    tt.cell(r,1).font=F(10,True,VIOLET)
    tt.cell(r,12).number_format="0%"
    tt.cell(r,9).number_format="yyyy-mm-dd"
    tt.cell(r,10).number_format="yyyy-mm-dd"
    tt.row_dimensions[r].height=30

# example row (T-001) — realistic values so the format is obvious
tt["G2"]="e.g. Dev name"; tt["I2"]=datetime.date(2026,8,4); tt["J2"]=datetime.date(2026,8,5)
tt["K2"]="In Progress"; tt["L2"]=0.6
tt["G2"].font=F(10,it=True,color=MUTED)
tt["A2"].comment=Comment("Example row — overwrite Assignee / dates / Status / % as work begins.","Tracker")

# data validation
def dv_add(col_letter, src, allow_blank=True):
    d=DataValidation(type="list",formula1=src,allow_blank=allow_blank)
    tt.add_data_validation(d); d.add(f"{col_letter}2:{col_letter}{LAST}")
dv_add("B", "Lists!$D$2:$D$8")
dv_add("E", "Lists!$C$2:$C$7")
dv_add("F", "Lists!$B$2:$B$4")
dv_add("K", "Lists!$A$2:$A$6")
pct=DataValidation(type="decimal",operator="between",formula1=0,formula2=1,allow_blank=True)
tt.add_data_validation(pct); pct.add(f"L2:L{LAST}")

# conditional formatting on Status
for val,c in [("Done",G_DONE),("In Progress",G_PROG),("Blocked",G_BLOCK),
              ("In Review",G_REV),("Not Started",G_NEW)]:
    tt.conditional_formatting.add(f"K2:K{LAST}",
        CellIsRule(operator="equal",formula=[f'"{val}"'],fill=fill(c)))
# overdue: Due < today and not Done -> red text
tt.conditional_formatting.add(f"J2:J{LAST}",
    FormulaRule(formula=[f'AND($J2<>"",$J2<TODAY(),$K2<>"Done")'],
                font=Font(name=FONT,size=10,color="C00000",bold=True)))

tt.freeze_panes="D2"
tt.auto_filter.ref=f"A1:M{LAST}"
tt.sheet_view.showGridLines=False

# ============================================================ MILESTONES
ms=wb.create_sheet("Milestones")
mheaders=["ID","Milestone","Target","Type","Payment %","Status","Actual date","Notes"]
mwidths=[7,34,10,9,11,13,13,34]
for c,(h,w) in enumerate(zip(mheaders,mwidths),1):
    cell=ms.cell(1,c,h); cell.font=F(10,True,WHITE); cell.fill=fill(BLUE)
    cell.alignment=center; cell.border=box
    ms.column_dimensions[get_column_letter(c)].width=w
ms.row_dimensions[1].height=24
mrows=[
 ("M1","Foundations live","End W2","Demo",0.15,"Pending"),
 ("M2","First device graded end to end","End W4","Demo",0.15,"Pending"),
 ("M3","Review & grading accepted","End W6","Demo",0.15,"Pending"),
 ("M4","Feature complete · code freeze","End W8","Gate",0.20,"Pending"),
 ("M5","UAT & accuracy signed off","End W10","Gate",0.15,"Pending"),
 ("M6","Production go-live & handover","End W12","Gate",0.15,"Pending"),
 ("—","Warranty end (final payment)","+90 days","—",0.05,"Pending"),
]
for r,(mid,name,tgt,typ,pay,st) in enumerate(mrows,2):
    vals=[mid,name,tgt,typ,pay,st,"",""]
    for c,v in enumerate(vals,1):
        cell=ms.cell(r,c,v); cell.border=box; cell.font=F(10)
        cell.alignment=center if c in(1,3,4,5,6) else lefttop
    ms.cell(r,1).font=F(10,True,BLUE); ms.cell(r,5).number_format="0%"
    ms.cell(r,7).number_format="yyyy-mm-dd"; ms.row_dimensions[r].height=24
# total payment check
tr=len(mrows)+2
ms.cell(tr,4,"Total").font=F(10,True); ms.cell(tr,4).alignment=Alignment(horizontal="right")
ms.cell(tr,5,f"=SUM(E2:E{tr-1})").font=F(10,True); ms.cell(tr,5).number_format="0%"
ms.cell(tr,5).border=Border(top=Side(style="thin",color=INK))
mdv=DataValidation(type="list",formula1='"Pending,On track,At risk,Done"',allow_blank=True)
ms.add_data_validation(mdv); mdv.add(f"F2:F{len(mrows)+1}")
ms.freeze_panes="A2"; ms.sheet_view.showGridLines=False

# ============================================================ DASHBOARD
db=wb.create_sheet("Dashboard"); wb.move_sheet("Dashboard",-(len(wb.sheetnames)-1))
db.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHIJ",[2,20,10,10,3,14,9,9,9,10]): db.column_dimensions[col].width=w
TT="'Task Tracker'"
IDR=f"{TT}!$A$2:$A${LAST}"; KR=f"{TT}!$K$2:$K${LAST}"; ER=f"{TT}!$E$2:$E${LAST}"
BR=f"{TT}!$B$2:$B${LAST}"; HR=f"{TT}!$H$2:$H${LAST}"; LR=f"{TT}!$L$2:$L${LAST}"

db.merge_cells("B2:J2")
db["B2"]="AI Device Inspector — Delivery Tracker"; db["B2"].font=F(18,True,VIOLET)
db.merge_cells("B3:J3")
db["B3"]="Live view of the developer backlog. Figures update automatically from the Task Tracker sheet."
db["B3"].font=F(10,it=True,color=MUTED)

# KPI tiles (B5..)
kpis=[("Total tasks",f'=COUNTIF({IDR},"?*")',VIOLET),
      ("Completed",f'=COUNTIF({KR},"Done")',CYAN),
      ("In progress",f'=COUNTIF({KR},"In Progress")',BLUE),
      ("In review",f'=COUNTIF({KR},"In Review")',"B7791F"),
      ("Blocked",f'=COUNTIF({KR},"Blocked")',"C00000"),
      ("Not started",f'=COUNTIF({KR},"Not Started")',MUTED)]
r0=5
for i,(lab,fml,col) in enumerate(kpis):
    r=r0+i
    db.cell(r,2,lab).font=F(10,False,INK); db.cell(r,2).alignment=Alignment(horizontal="left",vertical="center")
    c=db.cell(r,3,fml); c.font=F(14,True,col); c.alignment=center
    db.cell(r,2).border=bottom; db.cell(r,3).border=bottom
# progress
db.cell(r0+6,2,"% complete (effort)").font=F(10,True,INK)
pc=db.cell(r0+6,3,f'=IFERROR(SUMPRODUCT({HR},{LR})/SUM({HR}),0)'); pc.font=F(14,True,VIOLET); pc.number_format="0%"; pc.alignment=center
db.cell(r0+7,2,"% complete (count)").font=F(10,True,INK)
pcc=db.cell(r0+7,3,f'=IFERROR(COUNTIF({KR},"Done")/COUNTIF({IDR},"?*"),0)'); pcc.font=F(12,True,MUTED); pcc.number_format="0%"; pcc.alignment=center

# By sprint table (F5..)
db.cell(4,6,"By sprint").font=F(11,True,INK)
sh=["Sprint","Total","Done","%"]
for c,h in enumerate(sh): db.cell(5,6+c,h).font=F(9,True,WHITE); db.cell(5,6+c).fill=fill(BLUE); db.cell(5,6+c).alignment=center
for i,sp in enumerate(SPRINTS):
    r=6+i
    db.cell(r,6,sp).font=F(10)
    db.cell(r,7,f'=COUNTIF({ER},"{sp}")').alignment=center
    db.cell(r,8,f'=COUNTIFS({ER},"{sp}",{KR},"Done")').alignment=center
    pcell=db.cell(r,9,f'=IFERROR(H{r}/G{r},0)'); pcell.number_format="0%"; pcell.alignment=center
    for c in range(6,10): db.cell(r,c).border=bottom;
    if db.cell(r,7).font.size is None: pass

# By workstream table (below sprint)
wr=6+len(SPRINTS)+1
db.cell(wr,6,"By workstream").font=F(11,True,INK)
for c,h in enumerate(["Workstream","Total","Done","%"]): db.cell(wr+1,6+c,h).font=F(9,True,WHITE); db.cell(wr+1,6+c).fill=fill(VIOLET); db.cell(wr+1,6+c).alignment=center
for i,ws in enumerate(WSS):
    r=wr+2+i
    db.cell(r,6,f"{ws} · {WS_NAME[ws]}").font=F(9)
    db.cell(r,7,f'=COUNTIF({BR},"{ws}")').alignment=center
    db.cell(r,8,f'=COUNTIFS({BR},"{ws}",{KR},"Done")').alignment=center
    pcell=db.cell(r,9,f'=IFERROR(H{r}/G{r},0)'); pcell.number_format="0%"; pcell.alignment=center
    for c in range(6,10): db.cell(r,c).border=bottom

# helper status table for the chart (columns L/M, hidden from view)
db.cell(5,12,"Status"); db.cell(5,13,"Count")
for i,s in enumerate(STATUSES):
    db.cell(6+i,12,s); db.cell(6+i,13,f'=COUNTIF({KR},"{s}")')
db.column_dimensions["L"].hidden=True; db.column_dimensions["M"].hidden=True
# chart: tasks by status — placed below the KPI block, clear of the side tables
chart=BarChart(); chart.type="col"; chart.title="Tasks by status"; chart.height=6.8; chart.width=13
chart.legend=None
data=Reference(db,min_col=13,min_row=5,max_row=5+len(STATUSES))
cats=Reference(db,min_col=12,min_row=6,max_row=5+len(STATUSES))
chart.add_data(data,titles_from_data=True); chart.set_categories(cats)
chart.y_axis.majorGridlines=None
db.add_chart(chart,"B23")

# freeze & tidy
db.sheet_view.tabSelected=True
wb.active=wb.sheetnames.index("Dashboard")

# ============================================================ LEGEND on tracker
lg_row=LAST+2
tt.cell(lg_row,1,"HOW TO USE").font=F(9,True,VIOLET)
notes=[
 "Fill the shaded input columns as work proceeds: Assignee, Start, Due, Status, % Complete, Notes.",
 "Status & Priority & Sprint & Workstream are dropdowns. Status colours the row; overdue Due dates turn red.",
 "The Dashboard sheet totals everything automatically — no need to edit it.",
 "Add a new task by copying a row; keep the Task ID unique (T-064, …).",
]
for i,n in enumerate(notes):
    tt.cell(lg_row+1+i,1,"•  "+n).font=F(9,color=MUTED)
    tt.merge_cells(start_row=lg_row+1+i,start_column=1,end_row=lg_row+1+i,end_column=13)
# shade input columns lightly in header to hint they're editable
for c in (7,9,10,11,12,13):
    tt.cell(1,c).fill=fill(VIOLET)  # keep header violet; input hint via legend

wb.save("AI-Device-Inspector-Tracker.xlsx")
print("saved AI-Device-Inspector-Tracker.xlsx | tasks:",N,"| sheets:",wb.sheetnames)

# ---- self-check: compute expected dashboard values in Python ----
from collections import Counter
st=Counter(t for *_ ,t in [(x,) for x in ["Not Started"]*N])  # all start Not Started except example
statuses=["Not Started"]*N; statuses[0]="In Progress"
print("expected Total:",N,"Done:",statuses.count("Done"),
      "InProgress:",statuses.count("In Progress"),"NotStarted:",statuses.count("Not Started"))
for sp in SPRINTS:
    tot=sum(1 for t in tasks if t[4]==sp)
    print(f"  {sp:9s} total={tot}")
for ws in WSS:
    tot=sum(1 for t in tasks if t[1]==ws)
    print(f"  {ws:8s} total={tot}")
print("total est days:",sum(t[6] for t in tasks))
